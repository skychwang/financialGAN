import pdb
import sys
import openpyxl
import datetime
import numpy as np
import os
import pandas as pd

def order_aggregation_multiple_days():
	raw_orders = [file for file in os.listdir("RMD/") if file.startswith("PN_Order")]
	# pdb.set_trace()
	for raw_order in raw_orders:
		order_aggregation_one_day(raw_order, time_interval=100)
	# pdb.set_trace()


def order_aggregation_one_day(order_filename, time_interval=100):
	wb = openpyxl.load_workbook("RMD/" + order_filename)
	sheet = wb.worksheets[0]

	time_interval = time_interval #in milliseconds
	trading_interval = 6.5 * 3600 * 1000
	num_intervals = int(trading_interval / time_interval)
	#num_intervals = 200

	# TO-DO preprocessing to find min_max range 600 1200 900 1500
	buy_min = 600
	buy_max = 1200
	buy_vector = np.zeros((buy_max - buy_min, num_intervals))
	sell_min = 900
	sell_max = 1500
	sell_vector = np.zeros((sell_max - sell_min, num_intervals))

	parse_time = order_filename.split("_")[3].split(".")[0]
	month = int(parse_time[:2])
	date = int(parse_time[2:4])
	year = 2000 + int(parse_time[4:6])
	save_filename = "output/" + parse_time + "_" + str(time_interval) + ".json"

	time_start = datetime.datetime(year, month, date, 9, 30, 0, 000000)
	time_end = time_start + datetime.timedelta(microseconds = time_interval * 1000)

	i = 2
	index = 0
	#sheet.max_row
	while i <= sheet.max_row:
		if i%5000 == 0:
			print("processed {} lines".format(i))
		time_stamp = datetime.datetime.strptime(str(sheet["C" + str(i)].value), '%Y/%m/%d %H:%M:%S.%f')
		if time_stamp < time_start:
			i = i + 1
		elif time_start <= time_stamp < time_end:
			if sheet["P" + str(i)].value != 0:
				if sheet["G" + str(i)].value == 0 and buy_min <= sheet["Q" + str(i)].value * 100 < buy_max:
					buy_vector[int(sheet["Q" + str(i)].value * 100 - buy_min), index] = buy_vector[int(sheet["Q" + str(i)].value * 100 - buy_min), index] + \
					sheet["P" + str(i)].value
				elif sheet["G" + str(i)].value == 1 and sell_min <= sheet["Q" + str(i)].value * 100 < sell_max:
					sell_vector[int(sheet["Q" + str(i)].value * 100 - sell_min), index] = sell_vector[int(sheet["Q" + str(i)].value * 100 - sell_min), index] + \
					sheet["P" + str(i)].value
			i = i + 1
		else:
			index = index + 1
			time_start = time_end
			time_end = time_start + datetime.timedelta(microseconds = time_interval * 1000)
	# pdb.set_trace()
	save_orders_json(save_filename, buy_vector, sell_vector)

def save_orders_json(save_filename, buy_vector, sell_vector):
    order_dict = {"buy_vector": [buy_vector], "sell_vector": [sell_vector]}
    df = pd.DataFrame(data=order_dict, columns=["buy_vector", "sell_vector"])
    df.to_json(path_or_buf=save_filename, orient="records", lines=True)


