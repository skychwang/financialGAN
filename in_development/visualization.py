import matplotlib.pyplot as plt
from read_json import *
from order_vector import *
import operator
import scipy
import scipy.stats
from scipy.stats import poisson
import seaborn as sb
import math
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import numpy as np
from openpyxl import load_workbook


def get_size_distribution_per_price_level(orderstreams, is_buy=True, mode="average",
                                          sell_min=900, sell_max=1500,
                                          buy_min=600, buy_max=1200):
    #print(orderstreams.shape)
    num_of_batch = orderstreams.shape[0]

    num_batches = orderstreams.shape[0]
    batch_size = orderstreams.shape[1]
    order_stream_size = orderstreams.shape[2]
    if (is_buy):
        price_range = buy_max - buy_min
    else:
        price_range = sell_max - sell_min

    # convert the orderstream to a array of mutiple orders
    orderstreams = np.reshape(orderstreams, (num_batches * batch_size * order_stream_size, price_range, 2))

    buy_vector = orderstreams[:, :, 0]
    sell_vector = orderstreams[:, :, 1]
    #print(buy_vector.shape)
    #print(sell_vector.shape)

    # get x-axis
    if (is_buy):
        x = np.arange(buy_min, buy_max, 1.0)
        input_vector = buy_vector
    else:
        x = np.arange(sell_min, sell_max, 1.0)
        input_vector = sell_vector

    # get y-axis based on mode
    if (mode=="average"):
        #get the average size of each price
        y = np.mean(input_vector, axis=0)
    elif (mode=="median"):
        # get the median size of each price
        y = np.median(input_vector, axis=0)
    elif (mode=="sum"):
        # get the summation size of each price
        y = np.sum(input_vector, axis=0)
    elif (mode=="max"):
        #get the max size of each price
        y = np.amax(input_vector, axis=0)
    elif (mode=="min"):
        # get the min size of each price
        y = np.amin(input_vector, axis=0)
    elif (mode=="non_zero"):
        # get the min size of each price
        y = np.count_nonzero(input_vector, axis=0)
    elif (mode=="all"):
        # get the all size of each price
        fig = plt.figure()
        num_orders = input_vector.shape[0]

        x = []
        y = []
        for i in range(num_orders):
            if (is_buy):
                x += np.arange(buy_min, buy_max, 1).tolist()
            else:
                x += np.arange(sell_min, sell_max, 1).tolist()

            y += input_vector[i].tolist()

        plt.scatter(x, y, marker=".", s=1)

        if (is_buy):
            plt.xlabel("price(buy vector)")
        else:
            plt.xlabel("price(sell vector)")
        plt.ylabel(mode + " order size at this price")
        plt.legend(loc=1)
        if (is_buy):
            fig.savefig('(buy)all_size_distribution_per_price.png', dpi=200)
        else:
            fig.savefig('(sell)all_size_distribution_per_price.png', dpi=200)


        return

    # print(y)

    # print(np.amax(sell_vector))

    fig = plt.figure()


    plt.scatter(x, y, marker=".", s=1)
    if (is_buy):
        plt.xlabel("price(buy vector)")
    else:
        plt.xlabel("price(sell vector)")

    plt.ylabel(mode+" order size at this price")
    plt.legend(loc=1)
    plt.title(mode+' order size at each price level')
    if (is_buy):
        fig.savefig("(buy)"+mode + 'order_size_at_each_price_level.png', dpi=200)
    else:
        fig.savefig("(sell)"+mode + 'order_size_at_each_price_level.png', dpi=200)


def get_nonzero_element_distribution(orderstreams, is_buy=True, mode="nonzero"):
    num_of_batch = orderstreams.shape[0]
    num_batches = orderstreams.shape[0]
    batch_size = orderstreams.shape[1]
    order_stream_size = orderstreams.shape[2]

    orderstreams = np.reshape(orderstreams, (num_batches * batch_size * order_stream_size * 600, 2))
    buy_vector = orderstreams[:, 0]
    sell_vector = orderstreams[:, 1]

    if (is_buy):
        input_vector = buy_vector
    else:
        input_vector = sell_vector

    a = 1.01
    #input_vector = a - (a + 1) * np.exp(-(input_vector) ** 0.5 * np.log((a + 1) / (a - 1)))

    if (mode!="nonzero"):
        hist, bin_edges = np.histogram(input_vector, bins=int(np.amax(input_vector)))
        # print(hist)
        # print(bin_edges)
        num = len(input_vector)
        print("#######################################")
        print("the distribution of nonzero element are")
        for i in range(len(hist)):
            if (hist[i] != 0):
                print("size " + str(bin_edges[i]) + " show " + str(hist[i]) + " times" + \
                      ", which is " + str(hist[i] / num * 100) + "% of all")
    else:
        hist, bin_edges = np.histogram(input_vector, bins=np.arange(1,np.amax(input_vector),1))
        fig = plt.figure()
        plt.hist(input_vector, bins=np.arange(1,2000,1))
        plt.xlabel("size")
        plt.ylabel("frequency")
        plt.title("the hist of size freqency")
        if (is_buy):
            fig.savefig('(buy)the_hist_of_size_freqency.png', dpi=400)
        else:
            fig.savefig('(sell)the_hist_of_size_freqency.png.png', dpi=400)
        # print(hist)
        # print(bin_edges)
        num_nonzero = np.count_nonzero(input_vector)
        print("#######################################")
        print("the distribution of nonzero element are")
        for i in range(len(hist)):
            if (hist[i] != 0):
                print("size " + str(bin_edges[i]) + " show " + str(hist[i]) + " times" + \
                      ", which is " + str(hist[i] / num_nonzero * 100) + "% of all")



def get_nonzero_vector_distribution(orderstreams, is_buy=True, mode="hist",
                                    sell_min=900, sell_max=1500,
                                    buy_min=600, buy_max=1200):
    #print(orderstreams.shape)
    num_of_batch = orderstreams.shape[0]

    num_batches = orderstreams.shape[0]
    batch_size = orderstreams.shape[1]
    order_stream_size = orderstreams.shape[2]
    if (is_buy):
        price_range = buy_max - buy_min
    else:
        price_range = sell_max - sell_min

    # convert the orderstream to a array of mutiple orders
    orderstreams = np.reshape(orderstreams, (num_batches * batch_size * order_stream_size, price_range, 2))

    buy_vector = orderstreams[:, :, 0]
    sell_vector = orderstreams[:, :, 1]
    #print(buy_vector.shape)

    # get x-axis
    x = np.arange(0, orderstreams.shape[0], 1.0)
    if (is_buy):
        input_vector = buy_vector
    else:
        input_vector = sell_vector


    # get the number of nonzero element at each time interval(vector)
    y = np.count_nonzero(input_vector, axis=1)
    num_vectors = input_vector.shape[0]
    fig = plt.figure()
    if (mode == "hist"):
        plt.hist(y, bins=int(np.amax(y)))
        hist, bin_edges = np.histogram(y, bins=int(np.amax(y)))
        #print(hist)
        #print(bin_edges)
        print("#############################################################")
        print("the distribution of nonzero element in each time interval are")
        for i in range(len(hist)):
            if (hist[i] != 0):
                print(str(hist[i]) + " vectors has " + str(bin_edges[i]) + " nonzero element" + \
                      ", which is " + str(hist[i] / num_vectors*100) + "%")


    elif (mode=="hist_nonzero"):
        # get rid of all zeros
        #input_vector = input_vector[np.nonzero(input_vector)]
        #print(input_vector)
        #print(input_vector.shape)
        y = np.count_nonzero(input_vector, axis=1)

        plt.hist(y, bins=np.arange(1, np.amax(y),1))
        hist, bin_edges = np.histogram(y, bins=np.arange(1, np.amax(y),1))
        #print(hist)
        #print(bin_edges)
        print("#############################################################")
        print("the distribution of nonzero element in each time interval are")
        for i in range(len(hist)):
            if (hist[i]!=0):
                print(str(hist[i]) + " vectors has " + str(bin_edges[i]) + " nonzero element" + \
                      ", which is " + str(hist[i] / num_vectors*100) + "%")

    elif (mode == "scatter"):
        plt.plot(x, y, ".")
        plt.xlabel("order vectors(time intervel of 100ns)")

    elif (mode=="all"):
        # get the all size of each price
        fig = plt.figure()

        x = []
        y = []
        for i in range(price_range):
            x += np.arange(0, orderstreams.shape[0],1).tolist()
            temp = np.asarray(input_vector[:, i].tolist())
            temp[temp!=0] = i
            y += (temp.tolist())

        plt.scatter(x, y, marker=".", s=1)

        plt.xlabel("100ns time(each vector)")
        plt.ylabel("price distribution")
        plt.legend(loc=1)
        if (is_buy):
            fig.savefig('(buy)all_price_distribution_per_order.png', dpi=200)
        else:
            fig.savefig('(sell)all_price_distribution_per_order.png', dpi=200)


        return

    #print(y)
    plt.ylabel("number of nonzero element")
    plt.legend(loc=1)
    plt.title(mode + ' of nonzero distribution')
    if (is_buy):
        fig.savefig("(buy)"+mode + ' of nonzero distribution.png', dpi=200)
    else:
        fig.savefig("(sell)"+mode + ' of nonzero distribution.png', dpi=200)


# get the distribution of interval of two nonzero element
def get_time_interval_distribution(orderstreams, is_buy=True, mode="hist",
                                    sell_min=900, sell_max=1500,
                                    buy_min=600, buy_max=1200):
    #print(orderstreams.shape)
    num_of_batch = orderstreams.shape[0]

    num_batches = orderstreams.shape[0]
    batch_size = orderstreams.shape[1]
    order_stream_size = orderstreams.shape[2]
    if (is_buy):
        price_range = buy_max - buy_min
    else:
        price_range = sell_max - sell_min

    # convert the orderstream to a array of mutiple orders
    orderstreams = np.reshape(orderstreams, (num_batches * batch_size * order_stream_size, price_range, 2))

    buy_vector = orderstreams[:, :, 0]
    sell_vector = orderstreams[:, :, 1]
    #print(buy_vector.shape)

    # get x-axis
    x = np.arange(0, orderstreams.shape[0], 1.0)
    if (is_buy):
        input_vector = buy_vector
    else:
        input_vector = sell_vector

    # get a summation of the size varies with time
    y = np.sum(buy_vector, axis=1)

    #print(y)
    fig = plt.figure()

    if (mode=="hist"):
        last_nonzero = 0
        list = []
        # get the time interval of all nonzero element
        for i in range(1, y.shape[0]):
            if (y[i] != 0):
                list.append(i - last_nonzero - 1)
                last_nonzero = i

        #print(list)
        plt.hist(list, bins=np.arange(0, max(list), 1))

        hist, bin_edges = np.histogram(list, bins=np.arange(0, max(list), 1))
        plt.xlabel("number of 0 before nonzero vector appear")
        plt.ylabel("frequency")
        plt.legend(loc=1)
        plt.title('number of 0 before nonzero vector distribution')
        if (is_buy):
            fig.savefig('(buy)time_interval_distribution.png', dpi=200)
        else:
            fig.savefig('(sell)time_interval_distribution.png', dpi=200)
        print("###################################################################")
        print("the distribution of the time interval between nonzero vector are")

        for i in range(len(hist)):
            if (hist[i] != 0):
                print(str(hist[i]) + " nonzero vectors has " + str(bin_edges[i]) + " all zero vectors before it" + \
                      ", which is " + str(hist[i] / len(list)*100) + "%")
    else:
        #print(x.shape)
        #print(y[np.nonzero(y)].shape)
        plt.scatter(x, (y > 0), s=1)
        plt.show()




def get_distribution(orderstreams):
    # buy
    print("The distribution for buy vectors are")

    get_size_distribution_per_price_level(orderstreams, mode="max")

    get_size_distribution_per_price_level(orderstreams, mode="average")
    get_size_distribution_per_price_level(orderstreams, mode="non_zero")
    #get_size_distribution_per_price_level(orderstreams, mode="all")

    get_nonzero_element_distribution(orderstreams, mode="inclue_zero")
    get_nonzero_element_distribution(orderstreams)

    get_nonzero_vector_distribution(orderstreams, mode="hist")
    get_nonzero_vector_distribution(orderstreams, mode="hist_nonzero")
    # get_nonzero_vector_distribution(orderstreams, mode="all")

    get_time_interval_distribution(orderstreams, mode="hist")




    #sell
    print("The distribution for sell vectors are")
    get_size_distribution_per_price_level(orderstreams,is_buy=False, mode="max")

    get_size_distribution_per_price_level(orderstreams,is_buy=False, mode="average")
    get_size_distribution_per_price_level(orderstreams,is_buy=False, mode="non_zero")
    # get_size_distribution_per_price_level(orderstreams,is_buy=False, mode="all")

    get_nonzero_element_distribution(orderstreams,is_buy=False, mode="inclue_zero")
    get_nonzero_element_distribution(orderstreams,is_buy=False)

    get_nonzero_vector_distribution(orderstreams,is_buy=False, mode="hist")
    get_nonzero_vector_distribution(orderstreams,is_buy=False, mode="hist_nonzero")
    # get_nonzero_vector_distribution(orderstreams,is_buy=False, mode="all")

    get_time_interval_distribution(orderstreams,is_buy=False, mode="hist")


#orderstreams should be (num_time_intervals, 2), return the first 2000
def test_zero_one_distribution(orderstreams, mode="buy"):
    #print(orderstreams.shape)
    total_time_intervals = orderstreams.shape[0]

    if (mode == "buy"):
        buy_vector = orderstreams[:, 0]
        input_vector = buy_vector
    elif (mode == "sell"):
        sell_vector = orderstreams[:, 1]
        input_vector = sell_vector

    #print("the percentage of nonzero are")
    #print(str((np.count_nonzero(input_vector) / total_time_intervals) * 100) + "%")

    y = input_vector

    last_nonzero = 0
    list = []
    # get the time interval of all nonzero element
    for i in range(1, total_time_intervals):
        if (y[i] != 0):
            list.append(i - last_nonzero - 1)
            last_nonzero = i

    if (list == []):
        return np.zeros(200), 0, np.zeros(200)
    fig = plt.figure()

    plt.hist(list, bins=np.arange(0, max(max(list),201), 1))

    hist, bin_edges = np.histogram(list, bins=np.arange(0, max(max(list),201), 1))
    plt.xlabel("number of 0 before nonzero vector appear")
    plt.ylabel("frequency")
    plt.legend(loc=1)
    plt.title('number of 0 before nonzero vector distribution')
    fig.savefig(mode+'fake_buy_0-1.png', dpi=200)

    #print("###################################################################")
    #print("the distribution of the time interval between nonzero vector are")

    result = []
    for i in range(len(hist)):
        if (bin_edges[i] < 200):
            result.append(hist[i])

        #if (hist[i] != 0):
            #print(str(hist[i]) + " nonzero vectors has " + str(bin_edges[i]) + " all zero vectors before it" + \
                  #", which is " + str(hist[i] / len(list) * 100) + "%")
    return np.array(result), (np.count_nonzero(input_vector) / total_time_intervals), np.array(result) / len(list)


def test_0_1_real(file_name):
    orderstreams = np.load(file_name, mmap_mode='r')
    # get_distribution(orderstreams)

    print(orderstreams.shape)

    num_batches = orderstreams.shape[0]
    batch_size = orderstreams.shape[1]
    order_stream_size = orderstreams.shape[2]

    total_time_intervals = num_batches * batch_size * order_stream_size

    orderstreams = np.reshape(orderstreams, (total_time_intervals, 2))

    print("The buy vector is")
    dist_buy, non_zero_p_buy, norm_result_buy = test_zero_one_distribution(orderstreams, "buy")

    sum_x = 0
    for i in range(len(dist_buy)):
        # i is the number of failures, i+1 is the number of trials
        x = i+1
        # frequency of x number of trials occurs
        frequency = dist_buy[i]
        sum_x += x*frequency

    print(sum(dist_buy)/sum_x)

    p_buy = sum(dist_buy)/sum_x
    fit_result_buy = fit_geo_dist(dist_buy,p_buy)
    print("-------------------------------------------------")
    print("-------------------------------------------------")
    print("-------------------------------------------------")
    print("The sell vector is")
    dist_sell, non_zero_p_sell, norm_result_sell = test_zero_one_distribution(orderstreams, "sell")
    sum_x = 0
    for i in range(len(dist_sell)):
        # i is the number of failures, i+1 is the number of trials
        x = i + 1
        # frequency of x number of trials occurs
        frequency = dist_sell[i]
        sum_x += x * frequency

    print(sum(dist_sell) / sum_x)

    p_sell = sum(dist_sell) / sum_x
    fit_result_sell = fit_geo_dist(dist_sell, p_sell)
    return non_zero_p_buy, fit_result_buy, norm_result_buy, non_zero_p_sell, fit_result_sell, norm_result_sell


def test_0_1_fake(file_name, part=None, is_cancel=False):
    orderstreams = np.load(file_name, mmap_mode='r')
    # get_distribution(orderstreams)


    print(orderstreams.shape)
    num_runs = orderstreams.shape[0]

    if (part == None):
        # get all data
        steps = orderstreams.shape[1]
    else:
        steps = part

    result_buy = []
    result_sell = []
    buy_nonzero_list = []
    sell_nonzero_list = []

    for i in range(num_runs):
        orderstream_run = orderstreams[i,:,:]
        print(sum(orderstream_run))
        orderstream = np.zeros((steps, 2))
        print("the run " + str(i))

        for i in range(steps):
            if (is_cancel):
                if (orderstream_run[i][2] <= 0.5):
                    orderstream[i][0] = 0
                else:
                    orderstream[i][0] = 1

                if (orderstream_run[i][3] <= 0.5):
                    orderstream[i][1] = 0
                else:
                    orderstream[i][1] = 1
            else:
                if (orderstream_run[i][0] <= 0.5):
                    orderstream[i][0] = 0
                else:
                    orderstream[i][0] = 1

                if (orderstream_run[i][1] <= 0.5):
                    orderstream[i][1] = 0
                else:
                    orderstream[i][1] = 1

        orderstream = orderstream.astype(int)
        result_for_buy, buy_nonzero, norm_result_buy = test_zero_one_distribution(orderstream, "buy")
        result_for_sell, sell_nonzero, norm_result_sell = test_zero_one_distribution(orderstream, "sell")
        result_buy.append(norm_result_buy)
        buy_nonzero_list.append(buy_nonzero)
        result_sell.append(norm_result_sell)
        sell_nonzero_list.append(sell_nonzero)

    print("-------------------------------------------------")
    print("-------------------------------------------------")
    print("-------------------------------------------------")
    print("The buy vector has "  + str(np.mean(np.asarray(buy_nonzero_list))) + "% nonzero")
    result_np = np.asarray(result_buy)
    print(len(result_buy[0]))
    print(len(result_buy[1]))

    print(result_np.shape)

    mean_buy = np.mean(result_np, axis=0)
    std = np.std(result_np, axis=0)

    print("=================================================")
    print("The average result is")

    for i in range(200):
        if (mean_buy[i] != 0):
            print(str(mean_buy[i]) + " (mean) nonzero vectors has " + str(i) + " all zero vectors before it" + \
                  ", which is " + str(mean_buy[i] / sum(mean_buy) * 100) + "%" + ", the std is " + str(std[i]))


    print("-------------------------------------------------")
    print("-------------------------------------------------")
    print("-------------------------------------------------")
    print("The sell vector has " + str(np.mean(np.asarray(sell_nonzero_list))) + "% nonzero")
    print(len(result_sell[1]))
    result_np_sell = np.array(result_sell)

    print(result_np_sell.shape)

    mean_sell = np.mean(result_np_sell, axis=0)

    std = np.std(result_np_sell, axis=0)

    print("=================================================")
    print("The average result is")

    for i in range(200):
        if (mean_sell[i] != 0):
            print(str(mean_sell[i]) + " (mean) nonzero vectors has " + str(i) + " all zero vectors before it" + \
                  ", which is " + str(mean_sell[i] / sum(mean_sell) * 100) + "%" + ", the std is " + str(std[i]))
    if (sum(mean_buy) == 0):
        norm_result_buy = np.zeros(200)
    else:
        norm_result_buy = mean_buy / (sum(mean_buy))
    if (sum(mean_buy) == 0):
        norm_result_sell = np.zeros(200)
    else:
        norm_result_sell = mean_sell / (sum(mean_sell))
    return norm_result_buy,np.mean(np.asarray(buy_nonzero_list)), norm_result_sell, np.mean(np.asarray(sell_nonzero_list))


def fit_geo_dist(dist,p):
    plt.figure()
    fig, ax = plt.subplots(1, 1)
    mean, var, skew, kurt = scipy.stats.geom.stats(p, moments='mvsk')
    x = np.arange(1,201)
    ax.plot(x, scipy.stats.geom.pmf(x, p), 'bo', ms=8, label='geom pmf')
    ax.vlines(x, 0, scipy.stats.geom.pmf(x, p), colors='b', lw=5, alpha=0.5)
    rv = scipy.stats.geom(p)
    ax.vlines(x, 0, rv.pmf(x), colors='k', linestyles='-', lw=1, label = 'frozen pmf')
    ax.legend(loc='best', frameon=False)
    r = scipy.stats.geom.rvs(p, size=1000)

    fit_pmf = rv.pmf(x)

    #plt.show()
    return fit_pmf

def write_to_excel(real_file, fake_file, cancel, col):
    # write buy and sell
    non_zero_p_buy, fit_result_buy, real_result_buy, non_zero_p_sell, fit_result_sell, real_result_sell = test_0_1_real(
        real_file)
    fake_result_buy, fake_nonzero_buy, fake_result_sell, fake_nonzero_sell = test_0_1_fake(fake_file,
                                                                                           is_cancel=cancel)

    if cancel:
        name = "cancel"
    else:
        name = ""

    # sell
    diff_fit = [abs(x) for x in (fit_result_sell - real_result_sell)]
    diff_fake = [abs(x) for x in (fake_result_sell - real_result_sell)]
    print(sum(diff_fit))
    print(sum(diff_fake))

    df = pd.DataFrame({'A-real': real_result_sell,
                       'B-fit': fit_result_sell,
                       'C-diff(fit-real)': diff_fit,
                       'D-sum(fit-real)': sum(diff_fit),
                       'E-generated': fake_result_sell,
                       'F-diff(generated-real)': diff_fake,
                       'G-sum(generated-real)': sum(diff_fake),
                       'H-real_nonzero': non_zero_p_sell,
                       'I-fake_nonzero': fake_nonzero_sell
                       })
    book = load_workbook('result.xlsx')
    writer = ExcelWriter('result.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, 'Sell_'+name+'_GG', index=True, startrow=1, startcol=col)
    writer.save()

    # buy
    diff_fit = [abs(x) for x in (fit_result_buy - real_result_buy)]
    diff_fake = [abs(x) for x in (fake_result_buy - real_result_buy)]
    print(sum(diff_fit))
    print(sum(diff_fake))

    df = pd.DataFrame({'A-real': real_result_buy,
                       'B-fit': fit_result_buy,
                       'C-diff(fit-real)': diff_fit,
                       'D-sum(fit-real)': sum(diff_fit),
                       'E-generated': fake_result_buy,
                       'F-diff(generated-real)': diff_fake,
                       'G-sum(generated-real)': sum(diff_fake),
                       'H-real_nonzero': non_zero_p_buy,
                       'I-fake_nonzero': fake_nonzero_buy
                       })
    book = load_workbook('result.xlsx')
    writer = ExcelWriter('result.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, 'Buy_'+name+'_GG', index=True, startrow=1, startcol=col)
    writer.save()

def plot_possion():
    orderstreams = np.load("NPY/080116_100.npy", mmap_mode='r')
    # get_distribution(orderstreams)

    print(orderstreams.shape)

    num_batches = orderstreams.shape[0]
    batch_size = orderstreams.shape[1]
    order_stream_size = orderstreams.shape[2]

    total_time_intervals = num_batches * batch_size * order_stream_size

    orderstreams = np.reshape(orderstreams, (total_time_intervals, 2))


    total_time_intervals = orderstreams.shape[0]

    buy_vector = orderstreams[:, 0]
    input_vector = buy_vector

    # print("the percentage of nonzero are")
    # print(str((np.count_nonzero(input_vector) / total_time_intervals) * 100) + "%")

    y = input_vector

    _lambda = np.mean(input_vector)

    hist, bin_edges = np.histogram(input_vector, bins=int(np.amax(input_vector)))
    num = len(input_vector)
    print("#######################################")
    print("the distribution of nonzero element are")
    for i in range(len(hist)):
        if (hist[i] != 0):
            print("size " + str(bin_edges[i]) + " show " + str(hist[i]) + " times" + \
                  ", which is " + str(hist[i] / num * 100) + "% of all")


    print(_lambda)
    rv = scipy.stats.poisson(_lambda)
    data_binom = poisson.rvs(mu=_lambda, size=100000)
    fit_pmf = rv.pmf(np.arange(0,10))
    print(fit_pmf*100)
    ax = sb.distplot(data_binom,
                     kde=True,
                     color='green',
                     hist_kws={"linewidth": 25, 'alpha': 1})
    ax.set(xlabel='Poisson', ylabel='Frequency')
    plt.show()

def get_size_distribution_per_price_level_predict(orderstreams, is_buy=True, mode="average",
                                                  sell_min=900, sell_max=1500,
                                                  buy_min=600, buy_max=1200):
    x = np.arange(sell_min, sell_max, 1.0)
    input_vector = orderstreams
    # get y-axis based on mode
    if (mode=="average"):
        #get the average size of each price
        y = np.mean(input_vector, axis=0)
    elif (mode=="median"):
        # get the median size of each price
        y = np.median(input_vector, axis=0)
    elif (mode=="sum"):
        # get the summation size of each price
        y = np.sum(input_vector, axis=0)
    elif (mode=="max"):
        #get the max size of each price
        y = np.amax(input_vector, axis=0)
    elif (mode=="min"):
        # get the min size of each price
        y = np.amin(input_vector, axis=0)
    elif (mode=="non_zero"):
        # get the min size of each price
        y = np.count_nonzero(input_vector, axis=0)
    elif (mode=="all"):
        # get the all size of each price
        fig = plt.figure()
        num_orders = input_vector.shape[0]

        x = []
        y = []
        for i in range(num_orders):
            if (is_buy):
                x += np.arange(buy_min, buy_max, 1).tolist()
            else:
                x += np.arange(sell_min, sell_max, 1).tolist()

            y += input_vector[i].tolist()

        plt.scatter(x, y, marker=".", s=1)

        if (is_buy):
            plt.xlabel("price(buy vector)")
        else:
            plt.xlabel("price(sell vector)")
        plt.ylabel(mode + " order size at this price")
        plt.legend(loc=1)
        if (is_buy):
            fig.savefig('(buy)all_size_distribution_per_price.png', dpi=200)
        else:
            fig.savefig('(sell)all_size_distribution_per_price.png', dpi=200)
        return

    # print(y)

    # print(np.amax(sell_vector))

    fig = plt.figure()


    plt.scatter(x, y, marker=".", s=1)
    if (is_buy):
        plt.xlabel("price(buy vector)")
    else:
        plt.xlabel("price(sell vector)")

    plt.ylabel(mode+" order size at this price")
    plt.legend(loc=1)
    plt.title(mode+' order size at each price level')
    if (is_buy):
        fig.savefig("(buy)"+mode + 'order_size_at_each_price_level.png', dpi=200)
    else:
        fig.savefig("(sell)"+mode + 'order_size_at_each_price_level.png', dpi=200)


def get_nonzero_vector_distribution_predict(orderstreams, mode="hist"):
    input_vector = orderstreams
    # get x-axis
    x = np.arange(0, orderstreams.shape[0], 1.0)


    # get the number of nonzero element at each time interval(vector)
    y = np.count_nonzero(input_vector, axis=1)
    num_vectors = input_vector.shape[0]
    fig = plt.figure()
    if (mode == "hist"):
        plt.hist(y, bins=int(np.amax(y)))
        hist, bin_edges = np.histogram(y, bins=int(np.amax(y)))
        #print(hist)
        #print(bin_edges)
        print("#############################################################")
        print("the distribution of nonzero element in each time interval are")
        for i in range(len(hist)):
            if (hist[i] != 0):
                print(str(hist[i]) + " vectors has " + str(bin_edges[i]) + " nonzero element" + \
                      ", which is " + str(hist[i] / num_vectors*100) + "%")


    elif (mode=="hist_nonzero"):
        # get rid of all zeros
        #input_vector = input_vector[np.nonzero(input_vector)]
        #print(input_vector)
        #print(input_vector.shape)
        y = np.count_nonzero(input_vector, axis=1)

        plt.hist(y, bins=np.arange(1, np.amax(y),1))
        hist, bin_edges = np.histogram(y, bins=np.arange(1, np.amax(y),1))
        #print(hist)
        #print(bin_edges)
        print("#############################################################")
        print("the distribution of nonzero element in each time interval are")
        for i in range(len(hist)):
            if (hist[i]!=0):
                print(str(hist[i]) + " vectors has " + str(bin_edges[i]) + " nonzero element" + \
                      ", which is " + str(hist[i] / num_vectors*100) + "%")



def get_distribution_predict(file_name, mode="buy"):
    #convert the shape of orderstream
    orderstreams = np.load(file_name, mmap_mode='r')
    # get_distribution(orderstreams)

    print(orderstreams.shape)
    num_runs = orderstreams.shape[0]
    steps = orderstreams.shape[1]

    result_buy = []
    result_sell = []
    buy_nonzero_list = []
    sell_nonzero_list = []

    for i in range(num_runs):
        orderstream_run = orderstreams[i,:,:]
        print(sum(orderstream_run))
        orderstream = np.zeros((steps, 600))
        print("the run " + str(i))

        for i in range(steps):
            if (mode=="buy"):
                for j in range(0,600):
                    orderstream[i][j] = math.floor(orderstream_run[i][j]/100) * 100
            elif (mode=="sell"):
                for j in range(600,1200):
                    orderstream[i][j-600] = math.floor(orderstream_run[i][j]/100) * 100
            elif (mode=="buy_cancel"):
                for j in range(1200,1800):
                    orderstream[i][j-1200] = math.floor(orderstream_run[i][j]/100) * 100
            else:
                for j in range(1800,2400):
                    orderstream[i][j-1800] = math.floor(orderstream_run[i][j]/100) * 100
            print(i)
        print(orderstream.shape)

        get_nonzero_vector_distribution_predict(orderstream, mode="hist")


    # buy
    print("The distribution for buy vectors are")

    get_size_distribution_per_price_level(orderstreams, mode="max")

    get_size_distribution_per_price_level(orderstreams, mode="average")
    get_size_distribution_per_price_level(orderstreams, mode="non_zero")
    #get_size_distribution_per_price_level(orderstreams, mode="all")

    get_nonzero_element_distribution(orderstreams, mode="inclue_zero")
    get_nonzero_element_distribution(orderstreams)


    #sell
    print("The distribution for sell vectors are")
    get_size_distribution_per_price_level(orderstreams,is_buy=False, mode="max")

    get_size_distribution_per_price_level(orderstreams,is_buy=False, mode="average")
    get_size_distribution_per_price_level(orderstreams,is_buy=False, mode="non_zero")
    # get_size_distribution_per_price_level(orderstreams,is_buy=False, mode="all")

    get_nonzero_element_distribution(orderstreams,is_buy=False, mode="inclue_zero")
    get_nonzero_element_distribution(orderstreams,is_buy=False)


if __name__ == '__main__':
    #write_to_excel("NPY_cancel/080116_100.npy", "PN/day1.npy", cancel=True, col=0)
    """
    write_to_excel("NPY/080116_100.npy", "PN/day1.npy", cancel=False, col=0)

    #all 0 problem
    #write_to_excel("NPY_cancel/080116_100.npy", "PN/day1.npy", cancel=True, col=0)

    write_to_excel("NPY/080216_100.npy", "PN/day2.npy", cancel=False, col=11)
    write_to_excel("NPY_cancel/080216_100.npy", "PN/day2.npy", cancel=True, col=11)

    write_to_excel("NPY/080416_100.npy", "PN/day4.npy", cancel=False, col=22)
    write_to_excel("NPY_cancel/080416_100.npy", "PN/day4.npy", cancel=True, col=22)

    write_to_excel("NPY/080516_100.npy", "PN/day5.npy", cancel=False, col=33)
    write_to_excel("NPY_cancel/080516_100.npy", "PN/day5.npy", cancel=True, col=33)
    

    # Google
    write_to_excel("Google_data/NPY/080117_100.npy", "GG/day1.npy", cancel=False, col=0)
    write_to_excel("Google_data/NPY/080117_100.npy", "GG/day1.npy", cancel=True, col=0)

    write_to_excel("Google_data/NPY/080217_100.npy", "GG/day2.npy", cancel=False, col=11)
    write_to_excel("Google_data/NPY_cancel/080217_100.npy", "GG/day2.npy", cancel=True, col=11)

    write_to_excel("Google_data/NPY/080317_100.npy", "GG/day3.npy", cancel=False, col=22)
    write_to_excel("Google_data/NPY_cancel/080317_100.npy", "GG/day3.npy", cancel=True, col=22)

    write_to_excel("Google_data/NPY/080417_100.npy", "GG/day4.npy", cancel=False, col=33)
    write_to_excel("Google_data/NPY_cancel/080417_100.npy", "GG/day4.npy", cancel=True, col=33)
    """
    orderstreams = np.load("predict.npy", mmap_mode='r')
    print(orderstreams.shape)
    get_distribution_predict("predict.npy", mode="buy")










